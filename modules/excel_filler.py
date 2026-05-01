from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import DEFAULT_SOLAR_COST_PER_KW
from modules.utils import OUTPUTS_DIR, create_output_copy, ensure_default_template, load_cell_mapping


def generate_output_filename(source_name: str) -> str:
    stem = Path(source_name).stem.replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}_solar_analysis_{timestamp}.xlsx"


def generate_csv_filename(source_name: str) -> str:
    stem = Path(source_name).stem.replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}_solar_analysis_{timestamp}.csv"


def fill_excel_template(
    extracted_data: dict[str, Any],
    source_name: str,
    template_path: Path | None = None,
) -> Path:
    template_path = template_path or ensure_default_template()
    mapping = load_cell_mapping()
    output_path = create_output_copy(template_path, generate_output_filename(source_name))

    workbook = load_workbook(output_path)
    sheet = workbook.active

    for field, cell in mapping.items():
        if field not in extracted_data:
            continue
        sheet[cell] = extracted_data[field]

    workbook.save(output_path)
    return output_path


def generate_csv_output(
    extracted_data: dict[str, Any],
    source_name: str,
    template_path: Path | None = None,
) -> Path:
    template_path = template_path or ensure_default_template()
    mapping = load_cell_mapping()
    workbook = load_workbook(template_path)
    sheet = workbook.active

    for field, cell in mapping.items():
        if field not in extracted_data:
            continue
        sheet[cell] = extracted_data[field]

    output_path = OUTPUTS_DIR / generate_csv_filename(source_name)
    with output_path.open("w", newline="", encoding="utf-8") as file_obj:
        writer = csv.writer(file_obj)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])

    return output_path


def build_solar_summary(extracted_data: dict[str, Any]) -> dict[str, float | None]:
    units = _to_float(extracted_data.get("units_consumed"))
    load_kw = _to_float(extracted_data.get("connected_load_kw"))
    bill_amount = _to_float(extracted_data.get("bill_amount"))

    bill_per_unit = None
    if units and units > 0 and bill_amount is not None:
        bill_per_unit = round(bill_amount / units, 2)

    estimated_monthly_solar_offset_units = round(units * 0.75, 2) if units is not None else None
    suggested_system_size_kw = None
    if estimated_monthly_solar_offset_units is not None:
        derived_size = estimated_monthly_solar_offset_units / 120
        if load_kw is not None:
            suggested_system_size_kw = round(max(load_kw, derived_size), 2)
        else:
            suggested_system_size_kw = round(derived_size, 2)
    elif load_kw is not None:
        suggested_system_size_kw = round(load_kw, 2)

    estimated_monthly_savings = None
    if estimated_monthly_solar_offset_units is not None and bill_per_unit is not None:
        estimated_monthly_savings = round(estimated_monthly_solar_offset_units * bill_per_unit, 2)

    estimated_annual_savings = (
        round(estimated_monthly_savings * 12, 2)
        if estimated_monthly_savings is not None
        else None
    )

    estimated_system_cost = (
        round(suggested_system_size_kw * DEFAULT_SOLAR_COST_PER_KW, 2)
        if suggested_system_size_kw is not None
        else None
    )

    estimated_roi_years = None
    if estimated_system_cost and estimated_annual_savings and estimated_annual_savings > 0:
        estimated_roi_years = round(estimated_system_cost / estimated_annual_savings, 2)

    return {
        "bill_per_unit": bill_per_unit,
        "estimated_monthly_solar_offset_units": estimated_monthly_solar_offset_units,
        "suggested_system_size_kw": suggested_system_size_kw,
        "estimated_monthly_savings": estimated_monthly_savings,
        "estimated_annual_savings": estimated_annual_savings,
        "estimated_system_cost": estimated_system_cost,
        "estimated_roi_years": estimated_roi_years,
    }


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
