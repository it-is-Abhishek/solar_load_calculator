from __future__ import annotations

import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from PIL import Image

from config import (
    APP_LOG_FILE,
    CELL_MAPPING_PATH,
    DEFAULT_FIELD_MAPPING,
    DEFAULT_TEMPLATE_PATH,
    HISTORY_FILE,
    LOGS_DIR,
    OUTPUTS_DIR,
    TEMPLATES_DIR,
    UPLOADS_DIR,
)


def ensure_directories() -> None:
    for path in [TEMPLATES_DIR, UPLOADS_DIR, OUTPUTS_DIR, LOGS_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def setup_logging() -> logging.Logger:
    ensure_directories()
    logger = logging.getLogger("solar_load_calculator")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        formatter = logging.Formatter(
            "%(asctime)s | %(levelname)s | %(name)s | %(message)s"
        )
        file_handler = logging.FileHandler(APP_LOG_FILE, encoding="utf-8")
        file_handler.setFormatter(formatter)
        stream_handler = logging.StreamHandler()
        stream_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.addHandler(stream_handler)
    return logger


def save_uploaded_file(uploaded_file) -> Path:
    ensure_directories()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    safe_name = uploaded_file.name.replace(" ", "_")
    output_path = UPLOADS_DIR / f"{timestamp}_{safe_name}"
    with output_path.open("wb") as file_obj:
        file_obj.write(uploaded_file.getbuffer())
    return output_path


def load_cell_mapping() -> dict[str, str]:
    ensure_directories()
    if not CELL_MAPPING_PATH.exists():
        with CELL_MAPPING_PATH.open("w", encoding="utf-8") as file_obj:
            json.dump(DEFAULT_FIELD_MAPPING, file_obj, indent=2)
    with CELL_MAPPING_PATH.open("r", encoding="utf-8") as file_obj:
        return json.load(file_obj)


def ensure_default_template() -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    ensure_directories()
    if DEFAULT_TEMPLATE_PATH.exists():
        return DEFAULT_TEMPLATE_PATH

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Solar Analysis"

    sheet["A1"] = "Energybae Solar Load Calculator Template"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="D9EAF7")

    headers = [
        "Field",
        "Input Value",
        "Calculated Preview",
    ]
    for column, header in zip(["A", "B", "C"], headers):
        sheet[f"{column}2"] = header
        sheet[f"{column}2"].font = Font(bold=True)
        sheet[f"{column}2"].fill = PatternFill("solid", fgColor="EAEAEA")

    field_rows = {
        3: "Customer Name",
        4: "Consumer Number",
        5: "Billing Month",
        6: "Bill Amount",
        7: "Units Consumed",
        8: "Connected Load (kW)",
        9: "Tariff Category",
        10: "Meter Number",
        11: "Due Date",
        12: "Current Reading",
        13: "Previous Reading",
    }
    for row, label in field_rows.items():
        sheet[f"A{row}"] = label

    sheet["C6"] = "=IFERROR(B6/B7,0)"
    sheet["C7"] = "=IFERROR(B7*0.8,0)"
    sheet["C8"] = "=IFERROR(B8*1.25,0)"
    sheet["C15"] = "Estimated Monthly Solar Offset"
    sheet["D15"] = "=IFERROR(B7*0.75,0)"
    sheet["C16"] = "Suggested System Size (kW)"
    sheet["D16"] = "=IFERROR(MAX(B8,D15/120),0)"

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 20

    workbook.save(DEFAULT_TEMPLATE_PATH)
    return DEFAULT_TEMPLATE_PATH


def create_output_copy(template_path: Path, output_name: str) -> Path:
    ensure_directories()
    output_path = OUTPUTS_DIR / output_name
    shutil.copy2(template_path, output_path)
    return output_path


def append_history(record: dict[str, Any]) -> None:
    ensure_directories()
    with HISTORY_FILE.open("a", encoding="utf-8") as file_obj:
        file_obj.write(json.dumps(record, ensure_ascii=False) + "\n")


def load_history(limit: int = 50) -> list[dict[str, Any]]:
    if not HISTORY_FILE.exists():
        return []
    rows: list[dict[str, Any]] = []
    with HISTORY_FILE.open("r", encoding="utf-8") as file_obj:
        for line in file_obj:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return list(reversed(rows[-limit:]))


def history_to_dataframe(limit: int = 50) -> pd.DataFrame:
    records = load_history(limit=limit)
    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def clear_history() -> None:
    ensure_directories()
    HISTORY_FILE.write_text("", encoding="utf-8")


def tail_log(lines: int = 150) -> str:
    if not APP_LOG_FILE.exists():
        return "No logs available yet."
    with APP_LOG_FILE.open("r", encoding="utf-8") as file_obj:
        content = file_obj.readlines()
    return "".join(content[-lines:])


def image_to_preview(path: Path, max_size: tuple[int, int] = (900, 900)) -> Image.Image:
    image = Image.open(path)
    image.thumbnail(max_size)
    return image


def serialize_for_json(data: dict[str, Any]) -> str:
    return json.dumps(data, indent=2, ensure_ascii=False, default=str)
